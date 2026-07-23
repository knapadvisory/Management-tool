#!/usr/bin/env python3
"""
Marketplace fee-document parser -> Excel register (batch).
Supports AMAZON and FLIPKART. Handles Tax Invoices and Credit Notes,
intra-state (CGST+SGST) and inter-state (IGST), multi-page PDFs.

Credit notes are stored NEGATIVE (they reduce the TDS base) regardless of how
the marketplace prints them: Amazon prints them negative, Flipkart prints them
positive -- this tool normalises both.

USAGE (drag a folder onto Run_Parser.bat, or):
    python amazon_invoice_parser.py  <input_folder>  [<output_xlsx>]

Every document is reconciled against its own printed total. Edit TDS_MAP /
KEYWORD_MAP to change the section mapping.
"""

import sys, os, re, glob, datetime
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VERSION = "2026-07-22a (Amazon + Flipkart + Myntra + Nykaa; Myntra CN sign fix + duplicate flag & highlight)"

# ------------------------------------------------------------------ CONFIG
TDS_MAP = {
    "996211": ("194H", 0.02, "Commission (retail trade) - 194H"),
    "996729": ("194C*", 0.02, "Storage/removal - disputed; some treat as 194-I @10%"),
    "996812": ("194C",  0.02, "Shipping / weight handling / SCM (logistics)"),
    "997158": ("REVIEW", 0.0, "Payment-gateway charges - TDS treatment debated (194H exemption may apply); classify by hand"),
    "998365": ("194C",  0.02, "Advertising/marketing services - 194C @2% vs 194J @10%: confirm firm view"),
    "998541": ("194C",  0.02, "Giftwrap / packaging support"),
    "998599": ("REVIEW", 0.02, "998599 mixed; see keyword map"),
}
KEYWORD_MAP = [
    ("payment charges",("REVIEW", 0.0, "Payment-gateway charges - TDS treatment debated (194H exemption may apply); classify by hand")),
    ("payment",        ("REVIEW", 0.0, "Payment-gateway charges - TDS treatment debated (194H exemption may apply); classify by hand")),
    ("referral",       ("194H", 0.02, "Commission")),
    ("commission",     ("194H", 0.02, "Commission")),
    ("scm",            ("194C", 0.02, "Supply-chain/logistics")),
    ("marketing",      ("194C", 0.02, "Marketing/advertising - 194C @2% vs 194J @10%: confirm firm view")),
    ("collection fee", ("194H", 0.02, "Collection fee - 194H (commission-type) vs 194C: confirm firm view")),
    ("closing fee",    ("194H", 0.02, "Closing fee - 194H (commission-type) vs 194C: confirm firm view")),
    ("listing fee",    ("194H", 0.02, "Listing fee - 194H (commission-type) vs 194C: confirm firm view")),
    ("fixed fee",      ("194H", 0.02, "Fixed fee - 194H (commission-type) vs 194C: confirm firm view")),
    ("technology",     ("194J", 0.10, "Technology fee - 194J @10% (technical service) vs 194C: confirm")),
    ("add-ons amount recovery", ("REVIEW", 0.0, "Pass-through recovery - TDS may NOT apply; classify by hand")),
    ("recovery",       ("REVIEW", 0.0, "Recovery/reimbursement - TDS may NOT apply; classify by hand")),
    ("advertis",       ("194C", 0.02, "Advertising - some treat as 194J @10%")),
    ("ad services",    ("194C", 0.02, "Ad/advertising services - 194C @2% vs 194J @10%: confirm firm view")),
    ("ad service",     ("194C", 0.02, "Ad/advertising services - 194C @2% vs 194J @10%: confirm firm view")),
    ("storage",        ("194C*", 0.02, "Storage - disputed; some treat as 194-I @10%")),
    ("weight handling",("194C", 0.02, "Logistics")),
    ("shipping",       ("194C", 0.02, "Logistics")),
    ("pick and pack",  ("194C", 0.02, "Logistics support")),
    ("giftwrap",       ("194C", 0.02, "Packaging support")),
    ("removal",        ("194C", 0.02, "Logistics support")),
]
DEFAULT = ("REVIEW", 0.0, "Unmapped - classify manually")
TAX_WORDS = {"sgst", "cgst", "igst"}
GSTIN_RE = r"\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]Z[A-Z0-9]"


def tds_for(sac, desc):
    d = (desc or "").lower().replace("_", " ")
    for kw, val in KEYWORD_MAP:
        if kw in d:
            return val
    return TDS_MAP.get(sac, DEFAULT)


def grab(pattern, text, default=""):
    m = re.search(pattern, text, re.IGNORECASE)
    return m.group(1).strip() if m else default


def parse_amt(cell):
    if cell is None:
        return None
    s = str(cell).replace("\n", " ")
    m = re.search(r"([\d,]+\.\d{2})", s)
    if not m:
        return None
    v = float(m.group(1).replace(",", ""))
    return -v if "-" in s else v


def clean(cell):
    return (str(cell).replace("\n", " ").strip()) if cell else ""


# ------------------------------------------------------------------ AMAZON
def parse_amazon(pdf):
    page1 = pdf.pages[0].extract_text() or ""
    full, tables = page1, []
    for pg in pdf.pages:
        full += "\n" + (pg.extract_text() or "")
        tables.extend(pg.extract_tables())
    doc_type = "Credit Note" if "Credit Note" in page1 else "Tax Invoice"
    flat = full.replace("\n", " ")
    h = {
        "marketplace": "Amazon", "doc_type": doc_type,
        "number": grab(r"(?:Credit Note|Invoice) Number:\s*(\S+)", page1),
        "orig": "", "irn": "",
        "date": grab(r"(?:Credit Note|Invoice) Date:\s*([\d/]+)", page1),
        "sup_gstin": grab(r"GST Tax Registration No:\s*(\S+)", page1),
        "rec_name": grab(r"Name:\s*(.+)", page1),
        "rec_gstin": grab(r"GSTIN:\s*(\S+)", page1),
        "pos": grab(r"Place of Supply:\s*(.+)", page1),
    }
    h["stated_total"] = parse_amt(
        grab(r"Total Invoice amount\s*(-?INR[\s\d,\.]+)", flat)
        or grab(r"Total:\s*(-?INR[\s\d,\.]+)", flat))

    lines, cur, last_sac = [], None, ""
    for tbl in tables:
        if not tbl:
            continue
        head = " ".join(clean(c) for c in tbl[0]).lower()
        if "amount" not in head or "fee amount" in head:
            continue
        for row in tbl[1:]:
            cells = [clean(c) for c in row]
            desc = cells[-3] if len(cells) >= 3 else ""
            amt = parse_amt(row[-1]) if row else None
            sac_here = next((c for c in cells if re.fullmatch(r"\d{6}", c)), "")
            if sac_here:
                last_sac = sac_here
            dl = desc.lower()
            # A GST row (SGST/CGST/IGST) belongs to a fee line, never a fee line
            # of its own — attach it to the current line, or (if the previous
            # line was already flushed at a table boundary) the last one added.
            # Otherwise a standalone "SGST" row leaks in as a fake fee, wrongly
            # inflating the TDS base.
            if dl in TAX_WORDS:
                target = cur if cur else (lines[-1] if lines else None)
                if target is not None:
                    target[dl] = amt or 0.0
                continue
            if amt is None or not desc or "total" in dl or dl in ("amount", "description of service"):
                continue
            if cur:
                lines.append(cur)
            cur = {"sac": last_sac, "desc": desc, "taxable": amt, "sgst": 0.0, "cgst": 0.0, "igst": 0.0}
        if cur:
            lines.append(cur); cur = None
    if cur:
        lines.append(cur)
    return h, lines


# ------------------------------------------------------------------ FLIPKART
def parse_flipkart(pdf):
    page1 = pdf.pages[0].extract_text() or ""
    full, tables = page1, []
    for pg in pdf.pages:
        full += "\n" + (pg.extract_text() or "")
        tables.extend(pg.extract_tables())
    flat = full.replace("\n", " ")
    doc_type = "Credit Note" if "CREDIT NOTE" in full.upper() else "Tax Invoice"
    gstins = re.findall(GSTIN_RE, flat)
    sup = next((g for g in gstins if "AACCF" in g), "")
    rec = next((g for g in gstins if "AACCF" not in g), "")
    # Flipkart doc numbers: credit notes start FKC, invoices/originals start FKR
    toks = re.findall(r"FK[A-Z]{3}\d{9,}", flat)
    if doc_type == "Credit Note":
        number = next((t for t in toks if t.startswith("FKC")), "")
        orig = next((t for t in toks if t.startswith("FKR")), "")
    else:
        number = next((t for t in toks if t.startswith("FKR")), "")
        orig = ""
    h = {
        "marketplace": "Flipkart", "doc_type": doc_type,
        "number": number, "orig": orig,
        "date": grab(r"(?:Credit Note|Invoice) Date:\s*([\d\-/]+)", flat),
        "sup_gstin": sup, "rec_gstin": rec,
        "rec_name": grab(r"Business Name:\s*(.+?)\s+Address:", flat) or grab(r"Business Name:\s*(.+)", flat),
        "pos": grab(r"Place of Supply/State Code:\s*([^\n]+?)\s+Tel", flat) or grab(r"Place of Supply/State Code:\s*(.+)", flat),
        "irn": grab(r"IRN:\s*([0-9a-f]{16,})", flat),
    }
    sign = -1.0 if doc_type == "Credit Note" else 1.0

    # flatten all rows, locate the fee header row, map columns by name
    rows = [ [clean(c) for c in r] for tbl in tables for r in tbl ]
    hdr_idx = next((i for i, r in enumerate(rows)
                    if any("service accounting" in c.lower() for c in r)), None)
    lines, stated = [], None
    if hdr_idx is not None:
        hdr = rows[hdr_idx]
        def col(sub):
            return next((i for i, c in enumerate(hdr) if sub in c.lower()), None)
        i_tax = col("net taxable")
        i_igst = col("igst amount")
        i_cgst = col("cgst amount")
        i_sgst = col("sgst amount")
        i_tot = next((i for i, c in enumerate(hdr) if c.lower().startswith("total")), None)
        for r in rows[hdr_idx + 1:]:
            if not any(r):
                continue
            first = r[0].strip().lower()
            if first.startswith("total"):
                stated = parse_amt(r[i_tot]) if i_tot is not None else None
                break
            sac = next((c for c in r if re.fullmatch(r"\d{6}", c)), "")
            if not sac:
                continue
            desc = r[1] if len(r) > 1 else ""
            tax = parse_amt(r[i_tax]) if i_tax is not None else None
            if tax is None:
                continue
            ln = {"sac": sac, "desc": desc, "taxable": tax * sign, "sgst": 0.0, "cgst": 0.0, "igst": 0.0}
            if i_igst is not None and parse_amt(r[i_igst]) is not None:
                ln["igst"] = parse_amt(r[i_igst]) * sign
            if i_cgst is not None and parse_amt(r[i_cgst]) is not None:
                ln["cgst"] = parse_amt(r[i_cgst]) * sign
            if i_sgst is not None and parse_amt(r[i_sgst]) is not None:
                ln["sgst"] = parse_amt(r[i_sgst]) * sign
            lines.append(ln)
    h["stated_total"] = (stated * sign) if stated is not None else None
    return h, lines


# ------------------------------------------------------------------ MYNTRA
def parse_myntra(pdf):
    full = ""
    for pg in pdf.pages:
        full += "\n" + (pg.extract_text() or "")
    flat = full.replace("\n", " ")
    doc_type = "Credit Note" if "credit note" in full.lower() else "Tax Invoice"
    # Myntra prints credit notes as positive amounts (like Flipkart); store them
    # negative so they reduce the TDS base, consistent with every other source.
    sign = -1.0 if doc_type == "Credit Note" else 1.0
    gstins = re.findall(GSTIN_RE, flat)
    # supplier (Myntra) is the most frequently repeated GSTIN; recipient is the other
    sup = max(set(gstins), key=gstins.count) if gstins else ""
    rec = next((g for g in gstins if g != sup), "")
    names = re.findall(r"([A-Z][A-Za-z ]+PRIVATE LIMITED)", flat)
    rec_name = next((n.strip() for n in names if "MYNTRA" not in n.upper()), "")
    h = {
        "marketplace": "Myntra", "doc_type": doc_type,
        "number": grab(r"Invoice No\s*:\s*(\S+)", full),
        "orig": grab(r"Original Invoice No\s*:\s*(\S+)", full),
        "date": grab(r"Invoice Date\s*:\s*([\d\-/]+)", full),
        "sup_gstin": sup, "rec_gstin": rec, "rec_name": rec_name,
        "pos": grab(r"Place of supply\s*:\s*(\S+)", full),
        "irn": grab(r"IRN\s*:\s*([0-9a-f]{16,})", full),
    }
    lines, buf, stated = [], [], None
    for raw in full.splitlines():
        line = raw.strip()
        amts = re.findall(r"Rs\s*([\d,]+\.\d{2})", line)
        m = re.match(r"^(\d{6})\b(.*)$", line)
        if m and len(amts) >= 5:
            sac = m.group(1)
            rem = re.sub(r"Rs\s*[\d,]+\.\d{2}", "", m.group(2))
            rem = re.sub(r"\b\d+\.\d\b", "", rem).strip()  # drop qty like 1.0
            desc = rem if rem else " ".join(buf).strip()
            desc = re.sub(r"\(INR\)|\bRs\b|Unit Price", "", desc)
            desc = re.sub(r"\s+", " ", desc).strip()
            v = [float(a.replace(",", "")) * sign for a in amts]
            lines.append({"sac": sac, "desc": desc, "taxable": v[-5],
                          "igst": v[-4], "cgst": v[-3], "sgst": v[-2]})
            buf = []
        elif line.lower().startswith("total") and amts:
            stated = float(amts[-1].replace(",", "")) * sign
            buf = []
        elif "HSN/SAC" in line or line.startswith(("Billed", "Shipped")):
            buf = []
        elif re.search(r"[A-Za-z]", line) and ":" not in line \
                and "%" not in line and len(line) <= 45 \
                and "(INR)" not in line and "Unit Price" not in line \
                and not line.startswith("Rs"):
            buf.append(line)
            buf = buf[-3:]
    h["stated_total"] = stated
    return h, lines


# ------------------------------------------------------------------ NYKAA
def parse_nykaa(pdf):
    full = ""
    for pg in pdf.pages:
        full += "\n" + (pg.extract_text() or "")
    flat = full.replace("\n", " ")
    doc_type = "Credit Note" if "credit note" in full.lower() else "Tax Invoice"
    sup_gstin = grab(r"Supplier GSTIN\s*:\s*(\S+)", full)
    gstins = re.findall(GSTIN_RE, flat)
    rec = next((g for g in gstins if g != sup_gstin), "")
    h = {
        "marketplace": "Nykaa", "doc_type": doc_type,
        "number": grab(r"Document Number\s*:\s*(\S+)", full),
        "orig": grab(r"Original Invoice Number\s*:\s*([A-Za-z0-9]+)", full),
        "date": grab(r"Document Date\s*:\s*([\d.]+)", full),
        "sup_gstin": sup_gstin, "rec_gstin": rec,
        "rec_name": grab(r"Buyer\(Bill To\)\s*:\s*(.+)", full).replace("_DL", "").strip(),
        "pos": grab(r"Place of Supply\s*:\s*(.+)", full),
        "irn": grab(r"IRN No\s*:\s*(\S+)", full),
    }
    # split line-items (body) from Summary
    si = full.find("Summary")
    body = full[:si] if si > 0 else full
    summ = full[si:] if si > 0 else ""
    # SACs in order from summary lines
    sacs = re.findall(r"(?m)^\s*(\d{6})\s+[\d,]+\.\d{2}\s", summ)
    amt = r"([\d,]+\.\d{2})\s+[\d.]+%\s+([\d,]+\.\d{2})\s+[\d.]+%\s+([\d,]+\.\d{2})\s+[\d.]+%\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})"
    data_re = re.compile(r"^\s*\d+\s+(\S+)\s+.*?" + amt + r"\s*$")
    lines, buf, k = [], "", 0
    for raw in body.splitlines():
        line = raw.strip()
        m = data_re.match(line)
        if m:
            sku = m.group(1)
            desc = buf if buf else sku
            sac = sacs[k] if k < len(sacs) else ""
            k += 1
            f = lambda s: float(s.replace(",", ""))
            lines.append({"sac": sac, "desc": desc, "taxable": f(m.group(2)),
                          "cgst": f(m.group(3)), "sgst": f(m.group(4)), "igst": f(m.group(5)),
                          "_total": f(m.group(6))})
            buf = ""
        elif re.search(r"[A-Za-z]", line) and ":" not in line and "%" not in line \
                and len(line) <= 40 and "Wallet" not in line and "HSN" not in line \
                and "Description" not in line and not line[0:1].isdigit():
            buf = line
    # reconcile against the printed Total column (independent of the 4 components)
    h["stated_total"] = sum(l.pop("_total") for l in lines) if lines else None
    return h, lines


# ------------------------------------------------------------------ DISPATCH
def parse_doc(path):
    with pdfplumber.open(path) as pdf:
        head_text = (pdf.pages[0].extract_text() or "")
        if "Flipkart" in head_text:
            return parse_flipkart(pdf)
        if "MYNTRA" in head_text.upper():
            return parse_myntra(pdf)
        if "Nykaa" in head_text:
            return parse_nykaa(pdf)
        return parse_amazon(pdf)


# ------------------------------------------------------------------ BUILD
def build(folder, out_path):
    pdfs = sorted(glob.glob(os.path.join(folder, "*.pdf")))
    if not pdfs:
        print("No PDFs found in", folder); return

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Fee Register"
    FONT = "Arial"
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    DUP_FILL = PatternFill("solid", fgColor="FCE4D6")   # amber — duplicate rows
    CHECK_FILL = PatternFill("solid", fgColor="FFF2CC") # yellow — didn't reconcile
    OK_FILL = PatternFill("solid", fgColor="E2EFDA")    # green — reconciled
    ERR_FILL = PatternFill("solid", fgColor="F8CBAD")   # red-ish — parse error
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Marketplace", "Doc Type", "Doc Number", "Original Ref #", "Doc Date",
               "Supplier GSTIN", "Recipient", "Recipient GSTIN", "Place of Supply", "IRN",
               "SAC Code", "Fee Description", "Taxable Value", "CGST", "SGST", "IGST",
               "Total GST", "Line Total", "TDS Section", "TDS Rate", "TDS Amount", "Note"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    r, recon = 2, []
    seen_numbers = {}  # doc number -> first file it appeared in (to flag duplicates)
    for p in pdfs:
        try:
            h, lines = parse_doc(p)
        except Exception as e:
            recon.append((os.path.basename(p), "PARSE ERROR", str(e))); continue
        # Flag a repeated invoice/credit-note number: the same document uploaded
        # twice would otherwise be counted twice in the TDS base.
        num = h["number"] or os.path.basename(p)
        dup_of = seen_numbers.get(num)
        seen_numbers.setdefault(num, os.path.basename(p))
        for ln in lines:
            sec, rate, note = tds_for(ln["sac"], ln["desc"])
            if dup_of:
                note = (note + " | " if note else "") + f"DUPLICATE of {dup_of}"
            vals = [h.get("marketplace",""), h["doc_type"], h["number"], h.get("orig",""), h["date"],
                    h["sup_gstin"], h["rec_name"], h["rec_gstin"], h["pos"], h.get("irn",""),
                    ln["sac"], ln["desc"], ln["taxable"], ln["cgst"], ln["sgst"], ln["igst"],
                    f"=N{r}+O{r}+P{r}", f"=M{r}+Q{r}", sec, rate, f"=M{r}*T{r}", note]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = Font(name=FONT, size=10); cell.border = border
                if dup_of:
                    cell.fill = DUP_FILL  # highlight duplicate lines amber
            r += 1
        parsed = sum(l["taxable"] + l["sgst"] + l["cgst"] + l["igst"] for l in lines)
        stated = h.get("stated_total") or 0
        reconciled = abs(parsed - stated) < 0.5
        detail = f"{h.get('marketplace','?')} {h['doc_type']}: parsed {parsed:.2f} vs printed {stated:.2f}"
        if dup_of:
            flag = "DUPLICATE"
            detail = f"Same number as {dup_of} — counted twice? " + detail
        else:
            flag = "OK" if reconciled else "CHECK"
        recon.append((num, flag, detail))

    last = r - 1
    ws.cell(row=r, column=12, value="GRAND TOTAL").font = Font(name=FONT, bold=True)
    for col in (13, 14, 15, 16, 17, 18, 21):
        L = get_column_letter(col)
        ws.cell(row=r, column=col, value=f"=SUM({L}2:{L}{last})").font = Font(name=FONT, bold=True)

    money = '#,##0.00;(#,##0.00)'
    for row in ws.iter_rows(min_row=2, max_row=r):
        for cell in row:
            cell.border = border
            if cell.column in (13, 14, 15, 16, 17, 18, 21):
                cell.number_format = money; cell.alignment = Alignment(horizontal="right")
            elif cell.column == 20:
                cell.number_format = '0.0%'; cell.alignment = Alignment(horizontal="center")
    widths = [11, 11, 17, 17, 11, 17, 24, 17, 16, 20, 9, 24, 12, 9, 9, 9, 11, 12, 11, 8, 11, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Reconciliation")
    ws2.append(["Parser version:", VERSION, ""])
    ws2.cell(row=1, column=1).font = Font(name=FONT, bold=True)
    ws2.append(["Document", "Status", "Detail"])
    for c in range(1, 4):
        ws2.cell(row=2, column=c).font = Font(name=FONT, bold=True)
    status_fill = {"OK": OK_FILL, "CHECK": CHECK_FILL, "DUPLICATE": DUP_FILL, "PARSE ERROR": ERR_FILL}
    for row in recon:
        ws2.append(list(row))
        rr = ws2.max_row
        fill = status_fill.get(row[1])
        if fill:
            for c in range(1, 4):
                ws2.cell(row=rr, column=c).fill = fill
        ws2.cell(row=rr, column=2).font = Font(name=FONT, bold=True)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 60

    wb.save(out_path)
    ok = sum(1 for _, s, _ in recon if s == "OK")
    print(f"Parser version: {VERSION}")
    print(f"Processed {len(pdfs)} doc(s): {ok} reconciled OK, {len(recon)-ok} to CHECK.")
    for row in recon:
        print("  ", row)
    print("->", out_path)
    return recon


if __name__ == "__main__":
    import json as _json
    json_mode = "--json" in sys.argv
    args = [a for a in sys.argv[1:] if a != "--json"]
    if len(args) == 1:
        folder = args[0]
        out = os.path.join(folder, f"Register_{datetime.date.today().isoformat()}.xlsx")
    elif len(args) == 2:
        folder, out = args[0], args[1]
    else:
        print(__doc__); sys.exit(1)
    recon = build(folder, out) or []
    if json_mode:
        rows = [{"doc": d, "status": s, "detail": dt} for (d, s, dt) in recon]
        print("RECON_JSON:" + _json.dumps({"version": VERSION, "rows": rows, "output": out}))
